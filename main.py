from heapq import nlargest

import openai
from flask import *
from flask_cors import CORS
from openpyxl import load_workbook

openai.api_type = "azure"
openai.api_base = "https://hackmee1-fc.openai.azure.com/"
openai.api_version = "2023-03-15-preview"
openai.api_key = "a2f35be0feb04e2187945ef2e7c03b0b"

app = Flask(__name__)
CORS(app)
prev_input=[]
pre_output=[]

gunit_system = '''
You have to act as an AI shopkeeper and play the role of system only whose job is to channelize user intent based on the conversation so we can get orders.
Following are the details of the shop you are handling - it is a shop selling two categories - "kurtis" for women gender and "t-shirts" for male gender.
Within kurtis for women, there are 4 product attributes - color, fabric, occasion, sleeve length. Within color, we have 5 options - blue, yellow, pink, white, red. Within fabric, we have 3 options - rayon, cotton, crepe. Within occasion, we have 3 options - daily, party, festive. Within sleeves, we have 3 options - short sleeves, long sleeves, three-quarter sleeves.
Within t-shirts for men - color, fabric, neck, fabric, sleeve length. Within color, we have 5 options - white, black, yellow, blue, red. Within fabric, we have 3 options - cotton, cotton blend and polyester. Within neck, we have 2 options - round and polo. Within pattern, we have 3 options - printed, solid and stripes. Within sleeve, we have two options - short sleeves, long sleeves.
Now coming to the customer journey, user will ask a search term or will come with random query, based on that you need to have follow up conversation where you will ask a question and give options to user to select from. Please make sure you do not self-generate user responses for any questions.
 Every time you receive a response from the user, you have to first understand the intent of the user. A user can have one of the following intents:
1. Search - This means the user is trying to find the relevant products basis their needs. Example: "show me red kurtis" or "mujhe lal kurti dikhayein".
2. Scroll up - This means the user wants to scroll the page up to look for products in the same page which he has already seen. Example: "Mujhe pichle products dikhayien" or "show me the previous products" or "go up"
3. Scroll down - This means the user wants to scroll the page down to look for more new products.Example: "Aur products dikhayien" or "show me more" or "go down"
4. Pick - This means the user has liked one of the products he is seeing from the app and wants to see the product description for the selected product. Example: "Select the second product" or "mujhe doosra product dikhayein"
5. Exit - This means the user wishes to exit the shop. Example: "I do not wish to shop" or "band karo ye sab"
6. Size selection - This means the user wishes to select size of the particular product incase not already selected. For both kurti and t-shirt category, there are 16 possible options - XXS, XS, S, M, L, XL, XXL, XXXL, 4XL, 5XL, 6XL, 7XL, 8XL, 9XL, 10XL, Free Size.
7. Address - This means the user wishes to provide address or give confirmation on the address. Example: "Choose second address" or "doosra address select karne"
Please note that Size Selection, Address will only come after the Intent has been "Pick" in the "past response"
Once the intent is clear, follow the following user journey. Your goal is to ask relevant follow-up questions to get responses which allow moving to the next intent state. Use the "past message" key and "past response" key from the user response to setup past context. Use "next_question" key in the JSON response to pass questions as a string everytime you need to ask something to the user.
1.If the Intent is "Search", update the Intent to "Search" in the JSON response and do not ask any question related to size the user is looking for. Update the json response updating all the fields inside the "attributes" key based on the user response. If the exact mapping is not available, map it to the nearest fields based on the what the shop has. Example - If user response says "cyan color", update the color field inside the "attribute" key to "blue" (nearest color which is available in the shop).Mandatorily ask followup questions to the user on "attributes" which are not specified by the users in the "past response". Send the JSON response for the user. Assume the next Intent as “Pick” and keep on asking questions & basis the user responses, keep sending the updated JSON. Do not self-generate any user response. Ask for user response for every follow-up question. Additional constraint is you can only change intent to "Pick" if two of the fields in the "attributes" key have non-null values.
 If the last 3 "past response" key has "intent" key as "Search" , change "intent" in the JSON response to "Exit" and send to backend. Strictly follow the guard rails you have to keep in mind - never ask anything related brand or brand preference, for topics like price and size question keep it open ended. Example - Are you looking for any specific brand - This is a totally wrong question as it is out meesho's context. Also, in case user says a size, automatically map it to one of the 16 sizes mentioned above. For the prices, JSON response attribute "price" should have "under", "over" or "between" along with the price value to map any range a user shares. Example, if a user says, "price should be less than 500", JSON attribute "price" should store "under 500"
2. If Intent is "Scroll Up", update the Intent to "Scroll Up". Assume the next Intent as "Pick" and ask the follow-up question accordingly. Send in the JSON response.
3. If Intent is "Scroll Down", update the Intent to "Scroll Down". Assume the next Intent as "Pick" and ask the follow-up question accordingly. Send in the JSON response.
4. If Intent is "Pick", update the Intent to "Pick". Update & send the value of int variable "product_index" by extracting number from the user response statement in JSON format. Assume the next Intent as "Size selection", ask the follow-up question accordingly. Send in the JSON response.
5. If Intent is "Size selection", update the Intent to "Size selection" and update the key 'size_selection' in the json response by mapping user response to one of the 16 size options listed above and putting that as "size_selection" and assume the next Intent state as "Address". Example, if the user says "small size kurti", update "size_selection" to "S". Send the JSON response. Assume the next Intent as "Address", ask the follow-up question accordingly.
6. If Intent is “Address”, update the Intent to “Address” and update the value of "address" key by extracting number from the user response statement in JSON format. Send the JSON address. Assume the next Intent as “Exit” and ask the follow-up question accordingly.
Strictly follow the guard rails you have to keep in mind - never ask anything related brand or brand preference, for topics like price question keep it open ended. Example - Are you looking for any specific brand - This is a totally wrong question as it is out meesho's context.
Your response should be strictly in json format even the follow up questions in conversation should be in json only don't give any other format in output. Here is the output format of json you have to stick to for the entire chat -
{
 "intent": "string" (should be one of the above defined intent),
 "attributes" (this is only filled during the "Search" Intent. Use follow-up questions while in "Search" Intent to fill the values in this array): {
 "category": "string", (category - either kurti or t-shirt)
 "color": "string", (color should be mapped to ones we have defined for each category in the shop. If not defined for a category, leave it null)
 "fabric": "string", (fabric should be mapped to ones we have defined for each category in the shop. If not defined for a category, leave it null))
 "occasion":"string",
 "sleeve type": "string",
 "neck": "string",
 "pattern": "string",
 "price": "string",
 },
 "next_question": "string", (this is the variable which is used to ask follow-up questions)
"product_index": "int", (this is only filled when in "Pick" Intent)
"size_selection" : "string",
"address" : "int"
}
'''
class Size:
    def __init__(self, id, label):
        self.id = id
        self.label = label

class Address:
    def __init__(self, id, title, phone, address):
        self.id = id
        self.title = title
        self.phone = phone
        self.address = address


class Product:
    def __init__(self, pid, name, category_name, image, price, rating, description, sizes):
        self.id = pid
        self.name = name
        self.category_name = category_name
        self.image = image
        self.price = price
        self.rating = rating
        self.description = description
        self.sizes = sizes

    def __dict__(self):
        return {
            'id': self.id,
            'name': self.name,
            'category_name': self.category_name,
            'image': self.image,
            'price': self.price,
            'rating': self.rating,
            'description': self.rating,
            'sizes': self.sizes
        }



white_list_attrs =["action"]
attribute_inverse_index = {'color': {'red': {-1}, 'yellow': {-1}},
                           'price': {200: {-1}, 100: {-1}},
                           'size':{""}
                           }

intent_inp = ""

product_description = {"-": {-1}}
product_name = {'-': {1}}
products = {
    '-': Product(-1, "Product Name", "Category", ["image.jpg"], 19.99, 4.5, "Product description", [Size(0,"S"), Size(1,"M"), Size(3,"L")])}

address = {-1, Address(-1, "", 122, "vsdhcvjhs")}
category = {"saree": {-1}, "kurti": {-1}, "suits": {-1}, "tshirt": {-1}}

# price make lower call too in getPLPList


product_dict = {}

class Service:

    def roundup(self, st):
        return  int(st)
        #return int(math.ceil(x / 100.0)) * 100


    def dummy(self,dict) :
        x = 1

    def searchDictionary(self,value,reverseIndex):
       # print("for attr")
        #print(value)
        #print(reverseIndex)
        pids = reverseIndex[value]
        #print(pids)
        return pids


    def searchDictionaryForPrice(self,value,reverseIndex):
        #print("for price")
        #print(value)
        #print(reverseIndex)
        pids = {-1}
        for price in range(1,int(value)):
            pids = pids.union(reverseIndex[price])
        #print(pids)
        return pids


    def loadAsLists(self,path):
        path = "/final_attributes_excel.xlsx"
        l_data = load_workbook(path)
        sheet = l_data.active
        c_row = sheet.max_row
        c_column = sheet.max_column
        out_data=[]
        dict={}
        for a in range(1, c_row + 1):
            row = []
            for b in range(1, c_column + 1):
                ob = sheet.cell(row=a, column=b)
                row.append(ob.value)

            out_data.append(row)
        return out_data




    def addPLP(self,row):
        if str(row[0]) in products:
            return
        if ',' in row[8] :
            sizesList = row[8].split(",")
        else :
            sizesList = ["S", "M", "L"]

        sizes =    [Size(i+1,v) for i,v in enumerate(sizesList)]
        image = row[2].split(",")
        product = Product(row[0],row[3], row[6], image, int(row[1]), 4.5, row[7], sizes)
        products[str(row[0])] = (product)

    def intializeDatabase(self) :
        #with open(r'/Users/adityaraj/Downloads/hackathon-voice/final_attributes_excel.xlsx') as file_obj:

        # Create reader object by passing the file
        # object to reader method
        #reader_obj = csv.reader(file_obj)
        reader_obj = self.loadAsLists(r'/final_attributes_excel.xlsx')

        # Iterate over each row in the csv
        # file using reader
        idToHeader = {}
        #non_attributre = ""

        firstline = True
        for rowWithCase in reader_obj:
            print("rowcase")
            #print(rowWithCase)
            row = [str(x).lower() for x in rowWithCase]
            if firstline == True :
                idToHeader = {v: k for v, k in enumerate(row)}
                firstline = False
            else :
                print("row is")
               #(row)
                if row[4] not in attribute_inverse_index.keys():
                    attribute_inverse_index[row[4]] = {row[4]:-1}
                if row[5] not in attribute_inverse_index[row[4]].keys() :
                    attribute_inverse_index[row[4]][row[5]] = {-1}
                attribute_inverse_index[row[4]][row[5]].add(row[0])
                if self.roundup(row[1])  not in  attribute_inverse_index['price'].keys() :
                    attribute_inverse_index['price'][self.roundup(row[1])] = {-1}
                attribute_inverse_index['price'][self.roundup(row[1])].add(row[0])
                if row[3] not in  product_name :
                    product_name[row[3]] = {-1}
                product_name[row[3]].add(row[0])
                if row[7] not in  product_description :
                    product_description[row[7]] = {-1}
                product_description[row[7]].add(row[0])
                category[row[6]].add(row[0])
                self.addPLP(row)
           # print(row)
        print("category are")
        print(category)


    def getPLPList(self,searchDict):
        size = 20
        attributeList = searchDict.keys()
        filledAttributeList = []
        productScore = {-1:-1}
        print("ready 1")
        for attr in attributeList:
            print("ready 2")
            if searchDict[attr] != 'null' and searchDict[attr] != None and not white_list_attrs.__contains__(attr) :
                print("Attribute")
                print(attr)
                globalpids = {-1}
                if attr != 'category' :
                    if attr == 'price':
                        globalpids = self.searchDictionaryForPrice(searchDict[attr],attribute_inverse_index[attr])
                    else :
                        globalpids = self.searchDictionary(searchDict[attr],attribute_inverse_index[attr])

                print("result from reverse index")
                #print(globalpids)
                # do regex search here
                regexPids = {-1}
                if globalpids is None:
                    globalpids = {-1}
                for des,pids in product_description.items() :
                    if searchDict[attr] in des:
                        regexPids.union(pids)
                print("result from product_description regexx")
                #print(regexPids)
                for des,pids in product_name.items() :
                    if searchDict[attr] in des:
                        regexPids = regexPids.union(pids)
                print("result from product_name regexx")
               # print(regexPids)
                globalpids = globalpids.union(regexPids)

                filteredpid = []
                print("ready 3")
                print("all pids")
                #print(globalpids)

                if attr == 'category' :
                    globalpids = globalpids.union(category[searchDict['category']])

                print("category")
                #print(category)
                for pid in globalpids :
                    if pid in category[searchDict['category']]:
                        filteredpid.append(pid)

                print("ready 4")
                #print(filteredpid)

                for pid in filteredpid :
                    if pid == -1:
                        continue

                    if pid != productScore:
                        productScore[pid] = 0
                    productScore[pid] = productScore[pid] + 1



                print("ready 5")



        #print(productScore)
        sortedList  = nlargest(size, productScore, key=productScore.get)
        if "-1" in sortedList:
            sortedList.remove("-1")
        if -1 in sortedList:
            sortedList.remove(-1)
        return sortedList

    def getPLPPojoList(self,searchDict):
        print(searchDict)
        print("DELETE-1")
        sortedList = self.getPLPList(searchDict)
        #sortedList.remove(-1)
        print("sorted list")
        #print(sortedList)
        plpList = [products[pid] for pid in sortedList ]
        return plpList

    def getPLPResponseFromPids(self,pids):
        productReponseList = []
        for k in pids:
            if k in products:
                productReponseList.append((products[k]))

        return productReponseList


def func2(sizes):
    product_dicts = []
    for product in sizes:
        product_dict = {
            "id": product.id,
            "label": product.label
        }
        product_dicts.append(product_dict)
    return product_dicts


def plp():
    p1 = Product(1, "pid1", "kurtas", "", 100, 4.5, "Descpriopn", [Size(1,"M")])
    p2 = Product(2, "pid2", "kurtas", "", 110, 4.6, "Descpriopn", [Size(1,"M")])
    p3 = Product(3, "pid3", "kurtas", "", 120, 4.7, "Descpriopn", [Size(1,"M")])
    p4 = Product(4, "pid1", "kurtas", "", 120, 4.7, "Descpriopn",[Size(1,"M")])
    product_list = [p1, p2, p3, p4]
    return product_list


def func1(products_list):
    product_dicts = []
    for product in products_list:
        product_dict = {
            "id": product.id,
            "name": product.name,
            "category_name": product.category_name,
            "image": product.image,
            "price": product.price,
            "rating": product.rating,
            "description": product.description,
            "sizes": func2(product.sizes)
        }
        product_dicts.append(product_dict)
    return product_dicts


def search(input_string):
    message = "Can you find product category as category , color , fabric, occasion, sleeve type ,price from the string " + input_string + " reply each in one lower case as json capital word as json if not present mark as None for python dictionary"
    print(message)
    response = openai.ChatCompletion.create(
        engine="TheGeneratorsGPT35-16k",
        messages=[
            {"role": "user", "content": message}
        ]
        ,
        temperature=0.3,
        max_tokens=4000,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None)
    x = response.choices[0].message.content
    return x

def findIndex(product_descriptions,input_message):
    message = str(product_descriptions) + " Which product is the user trying to select from above options when he says " + input_message+" Answer in integer only without any logic explanation"
    print(message)
    response = openai.ChatCompletion.create(
        engine="TheGeneratorsGPT35-16k",
        messages=[
            {"role": "user", "content": message}
        ]
        ,
        temperature=0.3,
        max_tokens=4000,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None)
    index = response.choices[0].message.content
    return index

def buildProductDescriptionBased(pdp):
    formed_string=""
    i=1
    for product in pdp:
        formed_string+" " + i +" - " + str(product.name) +" , "
        i=i+1



@app.route("/temp1", methods=['POST'])
def temp1(searchAttribute):
    searchAttribute = {"color": "red", "category": "kurti"}
    dataRespone = service.getPLPPojoList(searchAttribute)
    return func1(dataRespone)

@app.route("/plp", methods=['POST'])
def fetchPlpAPI():
    searchAttribute = {"color": "red", "category": "kurti"}
    products_list = service.getPLPPojoList(searchAttribute)
    new_obj = {
        "plp": func1(products_list)
    }
    return new_obj

@app.route("/pdp", methods=['POST'])
def fetchPdpAPI():
    data = request.get_json()
    product = products[str(data['ids'][0])]
    return {
        "id": product.id,
        "name": product.name,
        "category_name": product.category_name,
        "image": product.image,
        "price": product.price,
        "rating": product.rating,
        "description": product.description,
        "sizes": func2(product.sizes)
    }


def buildAttributesForSearch(response):

    keys = response.keys()
    output = {}
    if (keys.__contains__("attributes")):
        attributes = response["attributes"]
        if (attributes.keys().__contains__("category")):
            output["category"] = attributes["category"]

        if (attributes.keys().__contains__("color")):
            output["color"] = attributes["color"]

        if (attributes.keys().__contains__("fabric")):
            output["fabric"] = attributes["fabric"]

        if (attributes.keys().__contains__("occasion")):
            output["occasion"] = attributes["occasion"]

        if (attributes.keys().__contains__("sleeve type")):
            output["sleeve type"] = attributes["sleeve type"]

        if (attributes.keys().__contains__("neck")):
            output["neck"] = attributes["neck"]

        if (attributes.keys().__contains__("price")):
            output["price"] = attributes["price"]

        if (attributes.keys().__contains__("size")):
            output["size"] = attributes["size"]

    return output



@app.route("/", methods=['POST'])
def f1():
    json = request.get_json()
    input_string = json["input_string"]
    message = "Answer in single word. Categorise the statement  for shopping app user input " + input_string + " into the following search / buy / pick  / scroll down / scroll up  / exit "
    print("Input Message : " + input_string)
    response = openai.ChatCompletion.create(
        engine="TheGeneratorsGPT35-16k",
        messages=[
            {"role": "user", "content": message}
        #    {"role": "system", "content": "A shopping app where user is searching by voice. Search means user is trying to query for a product "},
         #   {"role": "system", "content":" Exit means he wants to end the conversation. Pick means he wants to select a particular one based on the index of the item or the description of the product . Buy means user likes it and  ready to take the product"}
        ]
        ,
        temperature=0.1,
        max_tokens=800,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None)
    response=response.choices[0].message.content
    response = response.replace("null", "None")
    response = response.replace("t-shirt", "tshirt")
    response = response.replace("t-shirts", "tshirt")
    response = response.replace("kurtis", "kurti")
    response = eval(response)
    user_intention = response["intent"]
    products_list = []
    new_obj ={}
    print("user intent is : " + user_intention)
    if ("search" in user_intention.lower()):
        print("Search flow started")
        extracted_attributes =  buildAttributesForSearch(response)
            #search(message)
        print("Attributes Extaction : " )
        print(extracted_attributes)
        products_list = service.getPLPPojoList(extracted_attributes)
        print(products_list)
        print("Search flow ended")
        prev_input.append(input_string)
        pre_output.append(extracted_attributes)
        new_obj = {
            "user_intent": user_intention.upper(),
            "plp": func1(products_list)
        }

    elif("pick" in user_intention.lower()):
        pids_for_pickup= json["pids"]
        print("pids" + str(pids_for_pickup))
        plpBasedOnId=service.getPLPResponseFromPids(pids_for_pickup)
        print(plpBasedOnId)
        text=buildProductDescriptionBased(plpBasedOnId)
        print("text : "  +str(text))
        chosen_product_index = findIndex(text,input_string)
        new_obj = {
            "user_intent": user_intention.upper(),
            "pdp": chosen_product_index
        }


    #clear the prev msgs
    elif("exit" in user_intention.lower()):
        prev_input.clear()
        pre_output.clear()
        new_obj = {
            "user_intent": user_intention.upper(),
        }

    else:
        new_obj = {
        "user_intent": user_intention.upper(),
    }

    return new_obj

@app.route("/new", methods=['POST'])
def newSearch():
    json = request.get_json()
    input_string = ""
    print("PAST")
    print(prev_input)
    print(pre_output)
    print("PAST")
    raw_input_String=json["input_string"]
    for i in range(len(pre_output)):
        input_string = input_string + "past message : " + str(prev_input[i]) +" \n"
        input_string = input_string + "past response : " + str(pre_output[i]) +" \n"

    if(len(pre_output))>0:
        input_string = input_string + "message : "+  json["input_string"]
    else:
        input_string = input_string +  json["input_string"]

    message = "Answer with the json only for user input " + input_string
    print("Input Message : " + input_string)
    response = openai.ChatCompletion.create(
    engine="TheGeneratorsGPT35-16k",
        messages=[
            {"role": "user", "content": message},
            {"role": "system", "content": gunit_system}]
        ,
    temperature=0.5,
    max_tokens=800,
    top_p=0.95,
    frequency_penalty=0,
    presence_penalty=0,
    stop=None)

    response=response.choices[0].message.content
    raw_response=response
    print("Response")
    print(response)
    response = response.replace("null", "None")
    response = response.replace("t-shirt", "tshirt")
    response = response.replace("t-shirts", "tshirt")
    response = response.replace("kurtis", "kurti")
    print("response 2")
    response = eval(response)
    user_intention = response["intent"]
    products_list = []
    new_obj ={}
    print("user intent is : " + user_intention)
    next_qsn=""
    size_selection=""
    address=""

    if(response.keys().__contains__("next_question")):
        next_qsn =response["next_question"]

    if(response.keys().__contains__("size_selection")):
        size_selection=response["size_selection"]

    if(response.keys().__contains__("address")):
        address=response["address"]


    if(user_intention is None or user_intention ==""):
        new_obj = {
            "user_intent": user_intention.upper(),
            "plp": func1(products_list),
        }


    if ("search" in user_intention.lower()):
        print("Search flow started")
        extracted_attributes =  buildAttributesForSearch(response)
        #search(message)
        print("Attributes Extaction : " )
        print(extracted_attributes)
        products_list = service.getPLPPojoList(extracted_attributes)
        ##print(products_list)
        print("Search flow ended")
        prev_input.append(raw_input_String)
        pre_output.append(raw_response)
       # if()
        new_obj = {
            "user_intent": user_intention.upper(),
            "plp": func1(products_list),
        }

    elif("pick" in user_intention.lower()):
        pids_for_pickup= json["pids"]
        print("pids" + str(pids_for_pickup))
        plpBasedOnId=service.getPLPResponseFromPids(pids_for_pickup)
        print(plpBasedOnId)
        text=buildProductDescriptionBased(plpBasedOnId)
        print("text : "  +str(text))
        chosen_product_index = findIndex(text,input_string)
        new_obj = {
            "user_intent": user_intention.upper(),
            "pdp": chosen_product_index,
        }


    #clear the prev msgs
    elif("exit" in user_intention.lower()):
        prev_input.clear()
        pre_output.clear()
        new_obj = {
            "user_intent": user_intention.upper(),
        }

    else:
        new_obj = {
            "user_intent": user_intention.upper(),
        }

    new_obj["size_selection"]=size_selection
    new_obj["next_qsn"]=next_qsn
    new_obj["address"]=address

    return new_obj



@app.route("/fetchPlpBasedOnId", methods=['POST'])
def fetchPlpBasedOnIdApi(pids):
    products_list=service.getPLPResponseFromPids(pids)
    new_obj = {
        "plp": func1(products_list)
    }
    return new_obj

if __name__ == '__main__':
    service = Service()
    service.intializeDatabase()
    app.run(port=7777)
